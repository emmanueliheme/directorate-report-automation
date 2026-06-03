import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

FILES = {
    "staff":    "Staff_List.xlsx",
    "leave":    "FY2026LeaveReport.xlsx",
    "training": "FY26TrainingReport.xlsx",
}
DIRECTORATE_COL      = "Directorate"
LEAVE_DEPT_COL       = "Department"
LEAVE_UTIL_COL       = "Leave_Utilization"
LEAVE_TAKEN_VAL      = "Taken"
LEAVE_NOT_TAKEN_VAL  = "Not Taken"
TRAINING_DEPT_COL    = "Department_Name"
TRAINING_STATUS_COL  = "Sub_Status"
TRAINING_DONE_VAL    = "Completed"
TRAINING_PENDING_VAL = "Not Completed"
OUTPUT_FOLDER        = "directorate_reports"

C_NAVY  = "1F4E79"; C_ROW_ALT = "EBF3FB"; C_GREEN = "C6EFCE"
C_RED   = "FFC7CE"; C_BLUE    = "D6E4F0"; C_TOTAL = "BDD7EE"
C_YELLOW = "FFF2CC"

def bdr():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(ws, row, n, bg=C_NAVY):
    for c in range(1, n+1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        cell.fill = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = bdr()
    ws.row_dimensions[row].height = 22

def banner(ws, title, sub, n):
    sp = get_column_letter(n)
    ws.merge_cells(f"A1:{sp}1")
    c = ws["A1"]
    c.value = title; c.font = Font(bold=True, size=12, color=C_NAVY, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26
    ws.merge_cells(f"A2:{sp}2")
    c2 = ws["A2"]
    c2.value = sub; c2.font = Font(italic=True, size=9, color="595959", name="Calibri")
    c2.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 15
    ws.row_dimensions[3].height = 5

def autofit(ws):
    for col in ws.columns:
        ltr = get_column_letter(col[0].column)
        ml = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[ltr].width = min(ml + 4, 45)

def raw_sheet(wb, name, df, dire, hl=None, gv=None, bv=None):
    ws = wb.create_sheet(name)
    if df.empty:
        ws["A1"].value = f"No records for: {dire}"; return ws
    n = len(df.columns)
    banner(ws, f"{name}   —   {dire}",
           f"Generated: {datetime.now().strftime('%d %b %Y  %I:%M %p')}     Records: {len(df):,}", n)
    for ci, col in enumerate(df.columns, 1):
        ws.cell(row=4, column=ci, value=col)
    hdr(ws, 4, n)
    for ri, row in enumerate(df.itertuples(index=False), 5):
        alt = ri % 2 == 0
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(name="Calibri", size=9)
            cell.border = bdr()
            cell.alignment = Alignment(vertical="center")
            cell.fill = PatternFill("solid", start_color=C_ROW_ALT if alt else "FFFFFF")
            cn = df.columns[ci-1]
            if hl and cn == hl:
                if gv and val == gv: cell.fill = PatternFill("solid", start_color=C_GREEN)
                elif bv and val == bv: cell.fill = PatternFill("solid", start_color=C_RED)
    ws.freeze_panes = "A5"; autofit(ws); return ws

def leave_agg(wb, df, dire):
    ws = wb.create_sheet("Leave Aggregation")
    if df.empty: ws["A1"].value = f"No leave data for: {dire}"; return ws
    if LEAVE_UTIL_COL not in df.columns:
        if "Days_Taken" in df.columns:
            df = df.copy()
            df[LEAVE_UTIL_COL] = df["Days_Taken"].apply(lambda x: LEAVE_TAKEN_VAL if x > 0 else LEAVE_NOT_TAKEN_VAL)
        else: ws["A1"].value = f"Column {LEAVE_UTIL_COL} not found"; return ws
    piv = df.groupby([LEAVE_DEPT_COL, LEAVE_UTIL_COL]).size().unstack(fill_value=0).reset_index()
    for v in [LEAVE_TAKEN_VAL, LEAVE_NOT_TAKEN_VAL]:
        if v not in piv.columns: piv[v] = 0
    piv = piv[[LEAVE_DEPT_COL, LEAVE_TAKEN_VAL, LEAVE_NOT_TAKEN_VAL]].copy()
    piv.columns = ["Department", "Taken", "Not Taken"]
    piv["Total Staff"] = piv["Taken"] + piv["Not Taken"]
    piv["Utilization %"] = ((piv["Taken"] / piv["Total Staff"].replace(0,1))*100).round(1).astype(str)+"%"
    tk = piv["Taken"].sum(); ta = piv["Total Staff"].sum()
    tot = pd.DataFrame([{"Department":"TOTAL","Taken":tk,"Not Taken":piv["Not Taken"].sum(),
                          "Total Staff":ta,"Utilization %":f"{round(tk/ta*100,1)}%" if ta else "0%"}])
    piv = pd.concat([piv, tot], ignore_index=True)
    n = len(piv.columns)
    banner(ws, f"Leave Aggregation   —   {dire}", "Leave Utilization by Department", n)
    for ci, col in enumerate(piv.columns, 1): ws.cell(row=4, column=ci, value=col)
    hdr(ws, 4, n)
    for ri, row in enumerate(piv.itertuples(index=False), 5):
        is_tot = row[0] == "TOTAL"
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(name="Calibri", size=10, bold=is_tot, color=C_NAVY if is_tot else "000000")
            cell.border = bdr(); cell.alignment = Alignment(horizontal="center", vertical="center")
            if is_tot: cell.fill = PatternFill("solid", start_color=C_TOTAL)
            elif ci == 2: cell.fill = PatternFill("solid", start_color=C_GREEN)
            elif ci == 3: cell.fill = PatternFill("solid", start_color=C_RED)
            elif ci == 4: cell.fill = PatternFill("solid", start_color=C_YELLOW)
            else: cell.fill = PatternFill("solid", start_color="FFFFFF")
    ws.freeze_panes = "A5"; autofit(ws); return ws

def train_agg(wb, df, dire):
    ws = wb.create_sheet("Training Aggregation")
    if df.empty: ws["A1"].value = f"No training data for: {dire}"; return ws
    if TRAINING_STATUS_COL not in df.columns or TRAINING_DEPT_COL not in df.columns:
        ws["A1"].value = f"Required columns not found in training file"; return ws
    piv = df.groupby([TRAINING_DEPT_COL, TRAINING_STATUS_COL]).size().unstack(fill_value=0).reset_index()
    for v in [TRAINING_DONE_VAL, TRAINING_PENDING_VAL]:
        if v not in piv.columns: piv[v] = 0
    piv = piv[[TRAINING_DEPT_COL, TRAINING_DONE_VAL, TRAINING_PENDING_VAL]].copy()
    piv.columns = ["Department", "Completed", "Not Completed"]
    piv["Total Staff"] = piv["Completed"] + piv["Not Completed"]
    piv["Completion %"] = ((piv["Completed"] / piv["Total Staff"].replace(0,1))*100).round(1).astype(str)+"%"
    dk = piv["Completed"].sum(); da = piv["Total Staff"].sum()
    tot = pd.DataFrame([{"Department":"TOTAL","Completed":dk,"Not Completed":piv["Not Completed"].sum(),
                          "Total Staff":da,"Completion %":f"{round(dk/da*100,1)}%" if da else "0%"}])
    piv = pd.concat([piv, tot], ignore_index=True)
    n = len(piv.columns)
    banner(ws, f"Training Aggregation   —   {dire}", "Training Completion by Department", n)
    for ci, col in enumerate(piv.columns, 1): ws.cell(row=4, column=ci, value=col)
    hdr(ws, 4, n)
    for ri, row in enumerate(piv.itertuples(index=False), 5):
        is_tot = row[0] == "TOTAL"
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(name="Calibri", size=10, bold=is_tot, color=C_NAVY if is_tot else "000000")
            cell.border = bdr(); cell.alignment = Alignment(horizontal="center", vertical="center")
            if is_tot: cell.fill = PatternFill("solid", start_color=C_TOTAL)
            elif ci == 2: cell.fill = PatternFill("solid", start_color=C_GREEN)
            elif ci == 3: cell.fill = PatternFill("solid", start_color=C_RED)
            elif ci == 4: cell.fill = PatternFill("solid", start_color=C_YELLOW)
            else: cell.fill = PatternFill("solid", start_color="FFFFFF")
    ws.freeze_panes = "A5"; autofit(ws); return ws

def main():
    print(f"\n{'═'*68}")
    print(f"   Directorate Report Automation  v2  —  5-Sheet Build")
    print(f"{'═'*68}\n")
    dfs = {}
    for key, path in FILES.items():
        if not os.path.exists(path):
            print(f"  [ERROR] File not found: {path}"); return
        dfs[key] = pd.read_excel(path)
        print(f"  Loaded   {path:<35}  ({len(dfs[key]):>5,} rows)")
    for key, df in dfs.items():
        if DIRECTORATE_COL not in df.columns:
            print(f"\n  [ERROR] '{DIRECTORATE_COL}' not in {FILES[key]}")
            print(f"  Columns found: {list(df.columns)}"); return
    all_dirs = set()
    for df in dfs.values(): all_dirs.update(df[DIRECTORATE_COL].dropna().unique())
    directorates = sorted(all_dirs)
    print(f"\n  Directorates: {len(directorates)}")
    for d in directorates: print(f"    • {d}")
    print(); os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    errors = []
    for i, d in enumerate(directorates, 1):
        try:
            sf  = dfs["staff"][dfs["staff"][DIRECTORATE_COL]    == d].reset_index(drop=True)
            lf  = dfs["leave"][dfs["leave"][DIRECTORATE_COL]    == d].reset_index(drop=True)
            tf  = dfs["training"][dfs["training"][DIRECTORATE_COL] == d].reset_index(drop=True)
            wb  = openpyxl.Workbook(); wb.remove(wb.active)
            raw_sheet(wb, "Attrition",           sf, d)
            raw_sheet(wb, "FY2026LeaveReport",   lf, d, hl=LEAVE_UTIL_COL, gv=LEAVE_TAKEN_VAL, bv=LEAVE_NOT_TAKEN_VAL)
            leave_agg(wb, lf, d)
            raw_sheet(wb, "FY26 Training Report", tf, d, hl=TRAINING_STATUS_COL, gv=TRAINING_DONE_VAL, bv=TRAINING_PENDING_VAL)
            train_agg(wb, tf, d)
            safe = d.replace("/","-").replace(" ","_")
            wb.save(os.path.join(OUTPUT_FOLDER, f"{safe}_Report.xlsx"))
            print(f"  [{i:02d}/{len(directorates)}] {d:<35} Staff:{len(sf):>3}  Leave:{len(lf):>3}  Training:{len(tf):>3}  ✓")
        except Exception as e:
            errors.append((d, str(e))); print(f"  [{i:02d}/{len(directorates)}] {d:<35} ERROR: {e}")
    print(f"\n{'═'*68}")
    if errors: print(f"  Finished with {len(errors)} error(s).")
    else: print(f"  All {len(directorates)} reports saved  →  ./{OUTPUT_FOLDER}/")
    print(f"{'═'*68}\n")

if __name__ == "__main__":
    main()
